# -*- coding: utf-8 -*-
"""
生成資料庫 Table Schema Excel 文件 (每個表格一個 sheet)
根據 models.py 中的資料庫結構生成類似 docs/Table Schema.xlsx 的文件
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# 定義樣式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
table_name_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
table_name_font = Font(bold=True, size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 定義所有表格的 schema
tables = [
    {
        "name": "users",
        "description": "用戶資料表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "用戶ID (自動遞增)"},
            {"name": "username", "type": "TEXT", "pk": "", "nn": "Y", "unique": "Y", "default": "", "fk": "", "description": "用戶名稱 (唯一)"},
            {"name": "password_hash", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "密碼雜湊值"},
            {"name": "email", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "電子郵件"},
            {"name": "role", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "student", "fk": "", "description": "用戶角色 (student/teacher/admin)"},
            {"name": "created_at", "type": "DATETIME", "pk": "", "nn": "", "unique": "", "default": "CURRENT_TIMESTAMP", "fk": "", "description": "建立時間"},
        ]
    },
    {
        "name": "progress",
        "description": "學習進度表",
        "columns": [
            {"name": "user_id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "", "fk": "users(id)", "description": "用戶ID"},
            {"name": "skill_id", "type": "TEXT", "pk": "Y", "nn": "Y", "unique": "", "default": "", "fk": "skills_info(skill_id)", "description": "技能ID"},
            {"name": "consecutive_correct", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "0", "fk": "", "description": "連續答對次數"},
            {"name": "consecutive_wrong", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "0", "fk": "", "description": "連續答錯次數"},
            {"name": "current_level", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "1", "fk": "", "description": "當前等級"},
            {"name": "questions_solved", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "0", "fk": "", "description": "已解題數"},
            {"name": "last_practiced", "type": "DATETIME", "pk": "", "nn": "", "unique": "", "default": "CURRENT_TIMESTAMP", "fk": "", "description": "最後練習時間"},
        ]
    },
    {
        "name": "skills_info",
        "description": "技能單元資訊表",
        "columns": [
            {"name": "skill_id", "type": "TEXT", "pk": "Y", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "技能ID (唯一識別碼)"},
            {"name": "skill_en_name", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "技能英文名稱"},
            {"name": "skill_ch_name", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "技能中文名稱"},
            {"name": "category", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "分類"},
            {"name": "description", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "技能描述"},
            {"name": "input_type", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "text", "fk": "", "description": "輸入類型 (text/graph)"},
            {"name": "gemini_prompt", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "Gemini AI 提示詞"},
            {"name": "consecutive_correct_required", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "10", "fk": "", "description": "晉級所需連續答對次數"},
            {"name": "is_active", "type": "BOOLEAN", "pk": "", "nn": "", "unique": "", "default": "TRUE", "fk": "", "description": "是否啟用"},
            {"name": "order_index", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "0", "fk": "", "description": "排序索引"},
            {"name": "suggested_prompt_1", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "建議提示詞1"},
            {"name": "suggested_prompt_2", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "建議提示詞2"},
            {"name": "suggested_prompt_3", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "建議提示詞3"},
        ]
    },
    {
        "name": "skill_curriculum",
        "description": "課程綱要對應表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "ID (自動遞增)"},
            {"name": "skill_id", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "skills_info(skill_id)", "description": "技能ID"},
            {"name": "curriculum", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "課綱類型 (general/vocational)"},
            {"name": "grade", "type": "INTEGER", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "年級 (10/11/12)"},
            {"name": "volume", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "冊別 (數A/數B/數C等)"},
            {"name": "chapter", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "章節"},
            {"name": "section", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "小節"},
            {"name": "paragraph", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "段落 (可選)"},
            {"name": "display_order", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "0", "fk": "", "description": "顯示順序"},
            {"name": "difficulty_level", "type": "INTEGER", "pk": "", "nn": "", "unique": "", "default": "1", "fk": "", "description": "難度等級"},
        ]
    },
    {
        "name": "skill_prerequisites",
        "description": "技能前置依賴關聯表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "ID (自動遞增)"},
            {"name": "skill_id", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "skills_info(skill_id)", "description": "技能ID"},
            {"name": "prerequisite_id", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "skills_info(skill_id)", "description": "前置技能ID"},
        ]
    },
    {
        "name": "classes",
        "description": "班級資料表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "班級ID (自動遞增)"},
            {"name": "name", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "班級名稱"},
            {"name": "teacher_id", "type": "INTEGER", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "users(id)", "description": "教師ID"},
            {"name": "class_code", "type": "TEXT", "pk": "", "nn": "Y", "unique": "Y", "default": "", "fk": "", "description": "班級代碼 (唯一)"},
            {"name": "created_at", "type": "DATETIME", "pk": "", "nn": "", "unique": "", "default": "CURRENT_TIMESTAMP", "fk": "", "description": "建立時間"},
        ]
    },
    {
        "name": "class_students",
        "description": "班級-學生關聯表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "ID (自動遞增)"},
            {"name": "class_id", "type": "INTEGER", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "classes(id)", "description": "班級ID"},
            {"name": "student_id", "type": "INTEGER", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "users(id)", "description": "學生ID"},
            {"name": "joined_at", "type": "DATETIME", "pk": "", "nn": "", "unique": "", "default": "CURRENT_TIMESTAMP", "fk": "", "description": "加入時間"},
        ]
    },
    {
        "name": "mistake_logs",
        "description": "錯誤記錄表",
        "columns": [
            {"name": "id", "type": "INTEGER", "pk": "Y", "nn": "Y", "unique": "", "default": "AUTO", "fk": "", "description": "ID (自動遞增)"},
            {"name": "user_id", "type": "INTEGER", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "users(id)", "description": "用戶ID"},
            {"name": "skill_id", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "skills_info(skill_id)", "description": "技能ID"},
            {"name": "question_content", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "題目內容"},
            {"name": "user_answer", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "學生答案"},
            {"name": "correct_answer", "type": "TEXT", "pk": "", "nn": "Y", "unique": "", "default": "", "fk": "", "description": "正確答案"},
            {"name": "error_type", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "錯誤類型 (計算錯誤/觀念錯誤/粗心/其他)"},
            {"name": "error_description", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "錯誤描述"},
            {"name": "improvement_suggestion", "type": "TEXT", "pk": "", "nn": "", "unique": "", "default": "", "fk": "", "description": "改進建議"},
            {"name": "created_at", "type": "DATETIME", "pk": "", "nn": "", "unique": "", "default": "CURRENT_TIMESTAMP", "fk": "", "description": "建立時間"},
        ]
    },
]

def create_sheet_for_table(wb, table, is_first=False):
    """為每個表格創建一個 sheet"""
    if is_first:
        ws = wb.active
        ws.title = table['name']
    else:
        ws = wb.create_sheet(title=table['name'])
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 25  # Column Name
    ws.column_dimensions['B'].width = 12  # Data Type
    ws.column_dimensions['C'].width = 8   # PK
    ws.column_dimensions['D'].width = 8   # NN
    ws.column_dimensions['E'].width = 10  # Unique
    ws.column_dimensions['F'].width = 15  # Default
    ws.column_dimensions['G'].width = 20  # FK
    ws.column_dimensions['H'].width = 50  # Description
    
    # 寫入表格標題
    ws.merge_cells('A1:H1')
    title_cell = ws.cell(row=1, column=1, value=f"{table['name']} - {table['description']}")
    title_cell.fill = table_name_fill
    title_cell.font = table_name_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.border = border
    
    # 寫入欄位標題
    headers = ["Column Name", "Data Type", "PK", "NN", "Unique", "Default", "Foreign Key", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 寫入欄位資料
    for row_idx, col_info in enumerate(table['columns'], 3):
        ws.cell(row=row_idx, column=1, value=col_info['name']).border = border
        ws.cell(row=row_idx, column=2, value=col_info['type']).border = border
        ws.cell(row=row_idx, column=3, value=col_info['pk']).border = border
        ws.cell(row=row_idx, column=4, value=col_info['nn']).border = border
        ws.cell(row=row_idx, column=5, value=col_info['unique']).border = border
        ws.cell(row=row_idx, column=6, value=col_info['default']).border = border
        ws.cell(row=row_idx, column=7, value=col_info['fk']).border = border
        ws.cell(row=row_idx, column=8, value=col_info['description']).border = border
        
        # 對齊
        for col in range(1, 9):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 凍結前兩行
    ws.freeze_panes = 'A3'
    
    return ws

# 創建 Excel 工作簿
wb = openpyxl.Workbook()

# 為每個表格創建 sheet
for idx, table in enumerate(tables):
    create_sheet_for_table(wb, table, is_first=(idx == 0))

# 儲存檔案
output_dir = r'E:\Python\Mathproject\docs'
import os
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
output_file = os.path.join(output_dir, 'Database_Schema.xlsx')
wb.save(output_file)

print(f"[SUCCESS] 資料庫 Schema 已生成!")
print(f"檔案位置: {output_file}")
print(f"\n包含的 Sheets (共{len(tables)}個):")
for i, table in enumerate(tables, 1):
    print(f"  {i}. {table['name']} - {table['description']} ({len(table['columns'])} 欄位)")
