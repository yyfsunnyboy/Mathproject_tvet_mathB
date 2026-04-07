# -*- coding: utf-8 -*-
# 快速修改:使用新檔名避免權限問題

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 讀取原始腳本的表格定義
exec(open(r'c:\Mathproject\generate_db_schema.py', 'r', encoding='utf-8').read().split('# 創建 Excel 工作簿')[0])

# 創建 Excel 工作簿
wb = openpyxl.Workbook()

# 為每個表格創建 sheet
for idx, table in enumerate(tables):
    create_sheet_for_table(wb, table, is_first=(idx == 0))

# 使用新檔名
output_file = r'C:\Mathproject\docs\Database_Schema_Split.xlsx'
wb.save(output_file)

print(f"[SUCCESS] 資料庫 Schema 已生成!")
print(f"檔案位置: {output_file}")
print(f"\n包含的 Sheets (共{len(tables)}個):")
for i, table in enumerate(tables, 1):
    print(f"  {i}. Sheet '{table['name']}' - {table['description']} ({len(table['columns'])} 欄位)")
print(f"\n提示: 如果要覆蓋原檔案,請先關閉 Database_Schema.xlsx,然後重新執行 generate_db_schema.py")
